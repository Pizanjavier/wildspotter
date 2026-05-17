"""Import Spanish municipalities from INE xlsx + CNIG shapefiles into PostGIS.

Populates the municipalities table with ~8,131 Spanish municipalities.

Usage:
    python -m legal.import_municipalities --ine /data/legal/ine_municipios.xlsx --shp /data/legal/cnig_municipios/SHP_ETRS89/recintos_municipales_inspire_peninbal_etrs89/
    python -m legal.import_municipalities --ine /data/legal/ine_municipios.xlsx  # INE only (no geometry)
"""

import argparse
import logging
import re
import unicodedata
from pathlib import Path

from utils import get_db_connection, setup_logging

logger = setup_logging("import-municipalities")

PROVINCIA_TO_CCAA = {
    "01": "pais_vasco", "02": "castilla_la_mancha", "03": "valencia",
    "04": "andalucia", "05": "castilla_y_leon", "06": "extremadura",
    "07": "baleares", "08": "cataluna", "09": "castilla_y_leon",
    "10": "extremadura", "11": "andalucia", "12": "valencia",
    "13": "castilla_la_mancha", "14": "andalucia", "15": "galicia",
    "16": "castilla_la_mancha", "17": "cataluna", "18": "andalucia",
    "19": "castilla_la_mancha", "20": "pais_vasco", "21": "andalucia",
    "22": "aragon", "23": "andalucia", "24": "castilla_y_leon",
    "25": "cataluna", "26": "la_rioja", "27": "galicia", "28": "madrid",
    "29": "andalucia", "30": "murcia", "31": "navarra", "32": "galicia",
    "33": "asturias", "34": "castilla_y_leon", "35": "canarias",
    "36": "galicia", "37": "castilla_y_leon", "38": "canarias",
    "39": "cantabria", "40": "castilla_y_leon", "41": "andalucia",
    "42": "castilla_y_leon", "43": "cataluna", "44": "aragon",
    "45": "castilla_la_mancha", "46": "valencia", "47": "castilla_y_leon",
    "48": "pais_vasco", "49": "castilla_y_leon", "50": "aragon",
    "51": "ceuta", "52": "melilla",
}

PROVINCIA_NAMES = {
    "01": "Araba/Álava", "02": "Albacete", "03": "Alicante/Alacant",
    "04": "Almería", "05": "Ávila", "06": "Badajoz", "07": "Illes Balears",
    "08": "Barcelona", "09": "Burgos", "10": "Cáceres", "11": "Cádiz",
    "12": "Castellón/Castelló", "13": "Ciudad Real", "14": "Córdoba",
    "15": "A Coruña", "16": "Cuenca", "17": "Girona", "18": "Granada",
    "19": "Guadalajara", "20": "Gipuzkoa", "21": "Huelva", "22": "Huesca",
    "23": "Jaén", "24": "León", "25": "Lleida", "26": "La Rioja",
    "27": "Lugo", "28": "Madrid", "29": "Málaga", "30": "Murcia",
    "31": "Navarra", "32": "Ourense", "33": "Asturias", "34": "Palencia",
    "35": "Las Palmas", "36": "Pontevedra", "37": "Salamanca",
    "38": "Santa Cruz de Tenerife", "39": "Cantabria", "40": "Segovia",
    "41": "Sevilla", "42": "Soria", "43": "Tarragona", "44": "Teruel",
    "45": "Toledo", "46": "Valencia/València", "47": "Valladolid",
    "48": "Bizkaia", "49": "Zamora", "50": "Zaragoza",
    "51": "Ceuta", "52": "Melilla",
}


def normalize_name(name: str) -> str:
    name = name.lower().strip()
    name = unicodedata.normalize("NFD", name)
    name = re.sub(r"[̀-ͯ]", "", name)
    name = re.sub(r"[^a-z0-9 ]", " ", name)
    name = re.sub(r"\s+", " ", name).strip()
    return name


def natcode_to_ine(natcode: str) -> str:
    """INSPIRE NATCODE (11 digits) -> 5-digit INE code (CPRO+CMUN)."""
    return natcode[4:6] + natcode[8:11]


def import_ine_xlsx(ine_path: str) -> None:
    """Parse the multi-sheet INE xlsx (one sheet per province, 52 sheets)."""
    path = Path(ine_path)
    if not path.exists():
        logger.error("INE file not found: %s", ine_path)
        return

    try:
        import openpyxl
    except ImportError:
        logger.error("openpyxl not installed. Run: pip install openpyxl")
        return

    wb = openpyxl.load_workbook(path, read_only=True)
    conn = get_db_connection()
    inserted = 0

    try:
        with conn.cursor() as cur:
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                rows = list(ws.iter_rows(values_only=True))

                # Row 0: title, Row 1: province name, Row 2: headers (CPRO, CMUN, DC, NOMBRE)
                if len(rows) < 4:
                    continue

                provincia_name = str(rows[1][0]).strip() if rows[1][0] else ""

                for row in rows[3:]:
                    if not row or len(row) < 4:
                        continue
                    if row[0] is None or row[1] is None:
                        continue

                    cpro = str(int(row[0])).zfill(2) if isinstance(row[0], (int, float)) else str(row[0]).strip().zfill(2)
                    cmun = str(int(row[1])).zfill(3) if isinstance(row[1], (int, float)) else str(row[1]).strip().zfill(3)
                    nombre = str(row[3]).strip() if row[3] else ""

                    if not nombre or len(cpro) != 2:
                        continue

                    ine_code = cpro + cmun
                    ccaa = PROVINCIA_TO_CCAA.get(cpro, "unknown")
                    provincia = provincia_name or PROVINCIA_NAMES.get(cpro, "")
                    nombre_norm = normalize_name(nombre)

                    cur.execute(
                        """
                        INSERT INTO municipalities (ine_code, nombre, nombre_normalized,
                                                    provincia, provincia_code, ccaa)
                        VALUES (%s, %s, %s, %s, %s, %s)
                        ON CONFLICT (ine_code) DO UPDATE
                        SET nombre = EXCLUDED.nombre,
                            nombre_normalized = EXCLUDED.nombre_normalized,
                            provincia = EXCLUDED.provincia,
                            ccaa = EXCLUDED.ccaa
                        """,
                        (ine_code, nombre, nombre_norm, provincia, cpro, ccaa),
                    )
                    inserted += 1

        conn.commit()
        logger.info("Imported %d municipalities from INE", inserted)
    finally:
        wb.close()
        conn.close()


def shape_to_wkt(shape) -> str:
    """Convert a pyshp shape to WKT MultiPolygon."""
    parts = list(shape.parts) + [len(shape.points)]
    rings = []
    for i in range(len(parts) - 1):
        ring_points = shape.points[parts[i]:parts[i + 1]]
        ring_wkt = ", ".join(f"{x} {y}" for x, y in ring_points)
        rings.append(f"({ring_wkt})")

    # Group rings into polygons: outer ring (CCW in shapefile = CW in screen)
    # followed by holes. For simplicity, treat each ring set as one polygon.
    if len(rings) == 1:
        return f"MULTIPOLYGON((({rings[0][1:-1]})))"

    polygon_wkt = "(" + ", ".join(rings) + ")"
    return f"MULTIPOLYGON({polygon_wkt})"


def import_cnig_shapefile(shp_dir: str, source_srid: int = 4258) -> None:
    """Import CNIG shapefile geometries and join to municipalities by INE code."""
    import shapefile

    shp_path = Path(shp_dir)
    shp_files = list(shp_path.glob("*.shp"))
    if not shp_files:
        logger.error("No .shp files found in %s", shp_dir)
        return

    shp_file = shp_files[0]
    logger.info("Importing shapefile: %s", shp_file)

    prj_file = shp_file.with_suffix(".prj")
    if prj_file.exists():
        prj_text = prj_file.read_text()
        if "ETRS_1989" in prj_text or "ETRS89" in prj_text:
            source_srid = 4258
            logger.info("Detected ETRS89 (SRID 4258)")
        elif "REGCAN95" in prj_text:
            source_srid = 4081
            logger.info("Detected REGCAN95 (SRID 4081)")

    sf = shapefile.Reader(str(shp_file), encoding="latin-1")
    fields = [f[0] for f in sf.fields[1:]]
    natcode_idx = fields.index("NATCODE")
    logger.info("Fields: %s, NATCODE at index %d", fields, natcode_idx)

    conn = get_db_connection()
    matched = 0
    unmatched = 0

    try:
        with conn.cursor() as cur:
            for sr in sf.iterShapeRecords():
                natcode = sr.record[natcode_idx]
                if not natcode or len(str(natcode)) < 11:
                    unmatched += 1
                    continue

                ine_code = natcode_to_ine(str(natcode))

                try:
                    wkt = shape_to_wkt(sr.shape)
                except Exception as e:
                    logger.warning("Failed to parse geometry for %s: %s", ine_code, e)
                    unmatched += 1
                    continue

                cur.execute(
                    """
                    UPDATE municipalities
                    SET geom = ST_Multi(ST_Transform(
                        ST_GeomFromText(%s, %s), 4326
                    ))
                    WHERE ine_code = %s
                    """,
                    (wkt, source_srid, ine_code),
                )
                if cur.rowcount > 0:
                    matched += 1
                else:
                    unmatched += 1

            conn.commit()

        logger.info("Geometry matched: %d, unmatched: %d", matched, unmatched)
    finally:
        sf.close()
        conn.close()


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Import Spanish municipalities")
    parser.add_argument("--ine", required=True, help="Path to INE municipios .xlsx")
    parser.add_argument("--shp", help="Path to CNIG shapefile directory")
    args = parser.parse_args()

    import_ine_xlsx(args.ine)
    if args.shp:
        import_cnig_shapefile(args.shp)
